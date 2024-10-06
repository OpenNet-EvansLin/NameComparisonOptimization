import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import io

class SimilarUtil:
    @staticmethod
    def get_spin_similarity_ratio(str1: str, str2: str) -> float:
        if not str1 or not str2:
            return 0
        if str1.lower() == str2.lower():
            return 1.0

        len_str1 = len(str1)
        spin_ratio = 0.0

        for i in range(len_str1):
            temp = str1[len_str1-i:] + str1[:len_str1-i]
            score = SimilarUtil.get_similarity_ratio(temp, str2)
            spin_ratio = max(score, spin_ratio)

        return spin_ratio

    @staticmethod
    def get_similarity_ratio(str1: str, str2: str) -> float:
        max_len = max(len(str1), len(str2))
        return 1 - SimilarUtil.compare(str1, str2) / max_len if max_len > 0 else 0

    @staticmethod
    def compare(str1: str, str2: str) -> int:
        n, m = len(str1), len(str2)
        if n == 0:
            return m
        if m == 0:
            return n

        d = [[0] * (m + 1) for _ in range(n + 1)]

        for i in range(n + 1):
            d[i][0] = i
        for j in range(m + 1):
            d[0][j] = j

        for i in range(1, n + 1):
            ch1 = str1[i - 1]
            for j in range(1, m + 1):
                ch2 = str2[j - 1]
                if ch1.lower() == ch2.lower():
                    temp = 0
                else:
                    temp = 1
                d[i][j] = min(d[i - 1][j] + 1, d[i][j - 1] + 1, d[i - 1][j - 1] + temp)

        return d[n][m]

class SimilarityNameService:
    def process_v1(self, source: str, target: str) -> float:
        source = str(source) if source is not None else ""
        target = str(target) if target is not None else ""
        return SimilarUtil.get_spin_similarity_ratio(
            re.sub(r'\s+', ' ', source),
            re.sub(r'\s+', ' ', target)
        )

def create_histogram(data, output_path):
    plt.figure(figsize=(10, 6))
    plt.hist(data, bins=20, edgecolor='black')
    plt.title('Name Similarity Distribution')
    plt.xlabel('Similarity Score')
    plt.ylabel('Frequency')
    plt.savefig(output_path)
    plt.close()

input_file = 'InvetigateOpay350UserPhone1.csv'
output_file = 'InvetigateOpay350UserPhone1OutPutView.xlsx'
similarity_service = SimilarityNameService()
df = pd.read_csv(input_file)
df = df.dropna(subset=['account_name'])
df['new_similarity'] = df.apply(lambda row: similarity_service.process_v1(row['account_name'], row['submittedFullName']), axis=1)
wb = Workbook()
ws = wb.active
ws.title = "Similarity Results"
headers = df.columns.tolist()
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
for row, data in enumerate(df.values, start=2):
    for col, value in enumerate(data, start=1):
        ws.cell(row=row, column=col, value=value)
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column_letter].width = adjusted_width

img_buffer = io.BytesIO()
create_histogram(df['new_similarity'], img_buffer)
img_buffer.seek(0)

img = Image(img_buffer)
ws.add_image(img, 'A' + str(len(df) + 5))

wb.save(output_file)

print(f"處理完成：'{output_file}'。")
print(f"原始行數：{len(pd.read_csv(input_file))}")
print(f"處理後行數：{len(df)}")